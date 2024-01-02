import pandas as pd
import argparse as ap
from datetime import date, timedelta, datetime
import pyexcel as pe
import openpyxl
from file_setup import get_dir, file_check
from statistics import obtain_data, visualize
from dataclasses import dataclass

@dataclass
class Entry:
    late: bool 
    importance: bool
    input_date: str
    mood: int
    description: str
    sprdsht_dir: str

def cli_parse() -> None: 
    parser = ap.ArgumentParser(description="A CLI program to track your daily mood.")

    parser.add_argument("-e", "--enter", action="store_true", help="Input your daily mood.")
    parser.add_argument("-l", "--late", action="store_true", help="Use if you are submitting past 11:59pm.")
    parser.add_argument("-i", "--important", action="store_true", help="Mark as an important day.")
    parser.add_argument("-d", "--enter-date", type=str, help="Enter data for a specific day (yyyymmdd).", default=None)
    parser.add_argument("-t", "--test", action="store_true", help="fuck you")

    args = parser.parse_args()

    # Command for testing, modify appropriately 
    # if you are testing the code.
    if args.test:
        sprdsht_dir = get_dir()
        file_check(sprdsht_dir)
        dates, moods = obtain_data(30, sprdsht_dir)
        visualize(dates, moods)

    if args.enter or args.enter_date:
        late = args.late
        importance = args.important
        input_date = args.enter_date
        mood = get_mood()
        desc = get_desc()
        sprdsht_dir = get_dir()
        file_check(sprdsht_dir)
        entry = Entry(args.late, args.important, args.enter_date, mood, desc, sprdsht_dir)
        write_mood(entry)
        if args.enter_date:
            sort_data(sprdsht_dir)

def sort_data(sprdsht_dir: str) -> None:
    wb = openpyxl.load_workbook(sprdsht_dir)
    ws = wb.active

    # Takes the data from the spreadsheet 
    excel_data = []
    for row in ws.iter_rows(min_row = 2, values_only = True):
        excel_data.append(row)

    excel_data.sort(key=lambda row:datetime.strptime(row[0], '%Y-%m-%d'))
    
    # Deletes old, unsorted data
    ws.delete_rows(2, ws.max_row-1)

    # Appends new, sorted data
    for row in excel_data:
        ws.append(row)

    wb.save(sprdsht_dir)

def get_desc() -> str:
    return input("Please input a short description of your mood throughout the day: ")

def get_mood() -> int:
    mood = input("Please input your mood, from 1-10: ")
    try:
        mood_int = int(mood)
    except:
        print("Please input an integer.")
        return get_mood()
    if int(mood) > 10:
        print("Please enter an appropriate number.")
        return get_mood()
    else:
        return mood_int

# Checks to see if data already exists for 
# selected date.
def exists(entry: Entry, ws) -> bool:
    if ws.max_row >= 2:
        for rows in range(1, ws.max_row):
            if ws.cell(row=rows, column=1).value == str(entry.input_date):
                return True
        return False

def determine_date(entry: Entry) -> Entry:
    if entry.input_date != None:
        entry.input_date = datetime.strptime(entry.input_date, '%Y%m%d').date()
    elif not entry.late:
        entry.input_date = date.today()
    else:
        entry.input_date = date.today() - timedelta(days = 1)
    return entry

def write_mood(entry: Entry) -> None:
    entry = determine_date(entry)

    wb = openpyxl.load_workbook(entry.sprdsht_dir)
    ws = wb.active
    data = [str(entry.input_date), entry.mood, int(entry.importance), entry.description]
    
    # Make sure that the date doesn't already
    # contain data.
    if exists(entry, ws):
        print("You've already entered data.")
        exit()

    ws.append(data)
    wb.save(entry.sprdsht_dir)

def main():
    cli_parse()

if __name__ == "__main__":
    main()
