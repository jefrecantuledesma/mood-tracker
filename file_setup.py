from os.path import expanduser, isfile
import pandas as pd
import pyexcel as pe
import openpyxl

def get_dir():
    if isfile(expanduser('~/.config/mood_tracker/conf')):
        conf_dir = expanduser('~/.config/mood_tracker/conf') 
        with open(conf_dir, "a+") as config:
            config.seek(0)
            sprdsht_dir = config.readline()
            sprdsht_dir = sprdsht_dir.replace("file_dir = ", "")
            sprdsht_dir = sprdsht_dir.replace("\n", "")
            sprdsht_dir = expanduser(sprdsht_dir)
            return sprdsht_dir
    else:
        sprdsht_dir = expanduser('~/Documents/mood_tracker.xlsx')
        return sprdsht_dir

def file_check(sprdsht_dir: str):
    wb = openpyxl.load_workbook(sprdsht_dir)
    ws = wb.active
    if ws.cell(row=1, column=1).value != "Date":
        ws["A1"] = "Date"
        ws["B1"] = "Mood"
        ws["C1"] = "Importance"
        ws["D1"] = "Description"
        wb.save(sprdsht_dir)

